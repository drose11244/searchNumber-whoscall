# coding = utf-8
import os
import sys
from appium import webdriver
from time import sleep
from openpyxl import load_workbook, Workbook
import math


# Read xlsx

# getfilePath = sys.argv[1]
# loadWorkbook = load_workbook(getfilePath)

loadWorkbook = load_workbook('test2.xlsx')
# loadWorkbook = load_workbook('phoneNumberTest.xlsx')
sheetnames = loadWorkbook.get_sheet_names()
readsheet = loadWorkbook.get_sheet_by_name(sheetnames[0])

# create xlsx and typing value.
workbook = Workbook()
writesheet = workbook.active
writesheet.title = 'Sheet1'


# Function

def consoleLog(string):
    print(str(string))
    sys.stdout.flush()


def errorStatus(status):
    print("error_"+str(status))


def processbar(barLength, amount):
    bar = str(barLength) + "%"
    total = str(amount)+"/"+str(readsheet.max_row)

    if (barLength >= 100):
        setbarLength = 100
        bar = str(setbarLength) + "%"
        print("done_")

    print("width_" + str(bar))
    print("amount_" + str(total))
    sys.stdout.flush()


def getCountBarLength(total):
    result = math.ceil(100 / total)
    return result


# Test Device
def PATH(p): return os.path.abspath(
    os.path.join(os.path.dirname(__file__), p)
)


desired_caps = {}
desired_caps['platformName'] = 'Android'
# Check your android version.
# desired_caps['platformVersion'] = '5.0'
desired_caps['platformVersion'] = '4.4.4'

# Chose your device real or emulator
# desired_caps['deviceName'] = 'emulator'

# Typing your phoen's devicename
# you can input $adb devices to get device name
# desired_caps['deviceName'] = 'YT911ME0E'
desired_caps['deviceName'] = '0123456789ABCDEF'

desired_caps['app'] = PATH(
    '../apps/whoscall-6.33.apk'
    # '../../apps/whoscall-6.33.apk'
    # '../apps/ApiDemos-debug.apk'
)

desired_caps['appPackage'] = 'gogolook.callgogolook2'
desired_caps['appActivity'] = '.main.MainActivity'
# desired_caps['unicodeKeyboard'] = True
# desired_caps['resetKeyboard'] = True
print("init...", end='')

dr = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)

dr.tap([(0, 753), (480, 792)], 100)


print("pass")
# consoleLog("init_")
sleep(1)
print("start button...", end='')
dr.find_element_by_id('gogolook.callgogolook2:id/tv_start').click()
print("pass")
sleep(1)

print("skip button...", end='')
dr.find_element_by_id('gogolook.callgogolook2:id/btn_skip').click()
print("pass")
sleep(5)

print("search button...", end='')
dr.find_element_by_id(
    'gogolook.callgogolook2:id/menu_call_log_toolbar_search').click()
print('pass')
sleep(2)


getValue = readsheet.cell(row=1, column=1).value
# print xlsx
# print(getValue)
# sleep(1)
# dr.find_element_by_id(
#     'gogolook.callgogolook2:id/et_search_keyword').send_keys(getValue)

# sleep(1)

dr.find_element_by_id(
    'gogolook.callgogolook2:id/et_search_keyword').send_keys(getValue)

sleep(1)

dr.find_element_by_id(
    'gogolook.callgogolook2:id/et_search_keyword').click()

sleep(1)

# dr.hide_keyboard()
# dr.press_keycode(66)
# dr.press_keycode(84)
# dr.find_element_by_name('Go').click()
dr.tap([(0, 753), (480, 792)], 100)

sleep(1)

# print("type Enter ...")
result = dr.find_element_by_id(
    'gogolook.callgogolook2:id/line_primary')
text = result.text

sleep(5)
# number = 'A'+str(i)
writesheet['A'] = getValue

# number = 'B'+str(i)
writesheet['B'] = text
# print(text)
workbook.save('test1.xlsx')

dr.find_element_by_id(
    'gogolook.callgogolook2:id/et_search_keyword').clear()
sleep(1)
