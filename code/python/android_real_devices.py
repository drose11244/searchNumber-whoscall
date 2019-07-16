# coding = utf-8
import os
import sys
from appium import webdriver
from time import sleep
from openpyxl import load_workbook, Workbook
import math
import random
import string

# Function


def consoleLog(string):
    print("test_"+str(string))
    sys.stdout.flush()


def errorStatus(status):
    print("error_" + str(status))
    dr.close_app()


def processbar(barLength, amount):

    if (barLength >= 100):
        setbarLength = 100
        bar = str(setbarLength) + "%"
        dr.close_app()
        print("done_")

    bar = str(barLength) + "%"
    total = str(amount)+"/"+str(readsheet.max_row)

    print("width_" + str(bar))
    print("amount_" + str(total))
    sys.stdout.flush()


def getCountBarLength(total):
    result = math.ceil(100 / total)
    return result


def randomString(stringLength=20):
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))

# Test Device


def PATH(p): return os.path.abspath(
    os.path.join(os.path.dirname(__file__), p)
)


def recognizing():
    try:

        processBarLength = 0
        processbar(4, 0)
        fileName = randomString(10)

        for i in range(1, readsheet.max_row+1):
            getValue = readsheet.cell(row=i, column=1).value
            # print xlsx
            # print(getValue)
            sleep(1)
            dr.find_element_by_id(
                'gogolook.callgogolook2:id/et_search_keyword').send_keys(getValue)
            # print("type number ...")
            sleep(1)
            dr.find_element_by_id(
                'gogolook.callgogolook2:id/et_search_keyword').click()
            # dr.press_keycode(66)
            dr.tap([(0, 753), (480, 792)], 100)
            sleep(1)

            # print("type Enter ...")
            result = dr.find_element_by_id(
                'gogolook.callgogolook2:id/line_primary')
            text = result.text

            sleep(5)

            # consoleLog(text)
            # consoleLog(getValue)

            # get phoneNumber
            writesheet.cell(row=i, column=1).value = getValue
            writesheet.cell(row=i, column=2).value = text

            resultFileName = str(fileName) + ".xlsx"

            workbook.save(filename='done/'+resultFileName)

            sleep(1)

            dr.find_element_by_id(
                'gogolook.callgogolook2:id/et_search_keyword').clear()
            sleep(1)

            processBarLength += getCountBarLength(readsheet.max_row)
            processbar(processBarLength, i)

    except:
        errorStatus("file")


def init():
    try:
        # print("pass")
        consoleLog("init_")
        sleep(1)
        # print("start button...", end='')
        dr.find_element_by_id('gogolook.callgogolook2:id/tv_start').click()
        # print("pass")
        sleep(1)

        # print("skip button...", end='')
        dr.find_element_by_id('gogolook.callgogolook2:id/btn_skip').click()
        # print("pass")
        sleep(5)

        # print("search button...", end='')
        dr.find_element_by_id(
            'gogolook.callgogolook2:id/menu_call_log_toolbar_search').click()
        # print('pass')
        sleep(2)
    except:
        # print("error_init")
        errorStatus("init")


def main():
    init()
    recognizing()


if __name__ == "__main__":
    try:

        # Read xlsx

        getfilePath = sys.argv[1]
        loadWorkbook = load_workbook(getfilePath)

        # loadWorkbook = load_workbook('phoneNumberTest.xlsx')
        sheetnames = loadWorkbook.get_sheet_names()
        readsheet = loadWorkbook.get_sheet_by_name(sheetnames[0])

        # create xlsx and typing value.
        workbook = Workbook()
        writesheet = workbook.active
        writesheet.title = 'Sheet1'

        desired_caps = {}
        desired_caps['platformName'] = 'Android'
        # Check your android version.
        # desired_caps['platformVersion'] = '5.0'
        desired_caps['platformVersion'] = '4.4.4'

        # Chose your device real or emulator
        # desired_caps['deviceName'] = 'emulator'

        # Typing your phoen's devicename
        # you can input $adb devices to get device name
        # desired_caps['deviceName'] = 'YT911ME0E'
        desired_caps['deviceName'] = '0123456789ABCDEF'

        desired_caps['app'] = PATH(
            '../../apps/whoscall-6.33.apk'
            # '../apps/ApiDemos-debug.apk'
        )

        desired_caps['appPackage'] = 'gogolook.callgogolook2'
        desired_caps['appActivity'] = '.main.MainActivity'
        # print("init...", end='')

        dr = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)

        main()

    except:
        errorStatus("device")
