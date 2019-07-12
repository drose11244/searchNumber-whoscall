#coding = utf-8
import os
import sys
from appium import webdriver
from time import sleep
from openpyxl import load_workbook, Workbook


getfilePath = sys.argv[1]

# Read xlsx
loadWorkbook = load_workbook(getfilePath)
sheetnames = loadWorkbook.get_sheet_names()
readsheet = loadWorkbook.get_sheet_by_name(sheetnames[0])

# create xlsx and typing value.
workbook = Workbook()
writesheet = workbook.active
writesheet.title = 'Sheet1'


# Function
def processbar(barLength, amount):
    bar = str(barLength) + "%"
    total = str(amount)+"/"+str(readsheet.max_row)
    print("width_" + str(bar))
    print("amount_"+str(total))
    sys.stdout.flush()


# Test Device
def PATH(p): return os.path.abspath(
    os.path.join(os.path.dirname(__file__), p)
)


desired_caps = {}
desired_caps['platformName'] = 'Android'
# Check your android version.
desired_caps['platformVersion'] = '5.0'

# Chose your device real or emulator
# desired_caps['deviceName'] = 'emulator'

# Typing your phoen's devicename
# you can input $adb devices to get device name
desired_caps['deviceName'] = 'YT911ME0E'

desired_caps['app'] = PATH(
    '../apps/whoscall-6.33.apk'
    # '../apps/ApiDemos-debug.apk'
)

desired_caps['appPackage'] = 'gogolook.callgogolook2'
desired_caps['appActivity'] = '.main.MainActivity'
# print("init...", end='')

dr = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)


def recognizing():
    processBarLength = 0
    for i in range(1, readsheet.max_row):
        getValue = readsheet.cell(row=i, column=1).value
        # print xlsx
        # print(getValue)
        sleep(1)
        dr.find_element_by_id(
            'gogolook.callgogolook2:id/et_search_keyword').send_keys(getValue)
        print("type number ...")
        sleep(1)
        dr.find_element_by_id(
            'gogolook.callgogolook2:id/et_search_keyword').click()
        dr.press_keycode(66)
        sleep(1)

        # print("type Enter ...")
        result = dr.find_element_by_id(
            'gogolook.callgogolook2:id/line_primary')
        text = result.text

        sleep(5)
        number = 'A'+str(i)
        writesheet[number] = getValue

        number = 'B'+str(i)
        writesheet[number] = text
        # print(text)
        workbook.save('test1.xlsx')
        sleep(1)

        dr.find_element_by_id(
            'gogolook.callgogolook2:id/et_search_keyword').clear()
        sleep(1)

        # len = len + 9
        processBarLength += 9

        processbar(processBarLength, i)


def main():
    # print("pass")
    sleep(1)
    # print("start button...", end='')
    dr.find_element_by_id('gogolook.callgogolook2:id/tv_start').click()
    # print("pass")
    sleep(1)

    # print("skip button...", end='')
    dr.find_element_by_id('gogolook.callgogolook2:id/btn_skip').click()
    # print("pass")

    sleep(5)

    # print("search button...", end='')
    dr.find_element_by_id(
        'gogolook.callgogolook2:id/menu_call_log_toolbar_search').click()
    # print('pass')
    sleep(1)
    recognizing()

    # TEST data
    # processbar(9, 1)
    # sleep(1)

    # processbar(27, 3)
    # sleep(1)

    # processbar(100, 12)
    # sleep(1)


if __name__ == "__main__":
    main()


# processbar(9, 1)
# time.sleep(2)
# # processbar(18, 2)
# # processbar(27, 3)
# # processbar(36, 4)
# processbar(100, 12)
# time.sleep(2)


# print("python: " + getfilePath)
# print("width_" + str("10%"))
# print("amount_"+str(3)+"/"+str(readsheet.max_row))
# sys.stdout.flush()
# time.sleep(2)

# print("width_" + str("50%"))
# print("amount_"+str(10)+"/"+str(readsheet.max_row))
# sys.stdout.flush()
# time.sleep(2)

# print(str(3)+"/"+str(readsheet.max_row))
# sys.stdout.flush()
# time.sleep(2)

# print("width_" + str("100%"))
# print("amount_" + str(12) + "/" + str(readsheet.max_row))
# print("done_")
# sys.stdout.flush()
# time.sleep(2)
